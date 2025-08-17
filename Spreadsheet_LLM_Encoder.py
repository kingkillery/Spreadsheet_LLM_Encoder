import os
import openpyxl
import json
import logging
from temp_helpers import (
    infer_cell_data_type,
    categorize_number_format,
    get_number_format_string,
    detect_semantic_type,
)
from collections import defaultdict
from openpyxl.utils import get_column_letter

import sys

logger = logging.getLogger(__name__)


def calculate_compression_ratio(original_tokens: int, compressed_tokens: int) -> float:
    """Return the compression ratio given original and compressed token counts."""
    if compressed_tokens == 0:
        return 0.0
    if original_tokens == 0:
        return 1.0
    return original_tokens / compressed_tokens


def spreadsheet_llm_encode(excel_path, output_path=None, k=2, vanilla=False):
    """
    Convert an Excel file to SpreadsheetLLM format or a vanilla markdown-like format.

    Args:
        excel_path (str): Path to the Excel file.
        output_path (str, optional): Path to save the output. Defaults to None.
        k (int, optional): Neighborhood distance for structural anchors. Defaults to 2.
        vanilla (bool, optional): If True, produce vanilla encoding instead of compressed.
                                Defaults to False.

    Returns:
        dict: The SpreadsheetLLM encoding of the Excel file.
    """
    if vanilla:
        return vanilla_encode(excel_path, output_path)
    logger.info(f"Processing Excel file: {excel_path}")

    try:
        # openpyxl is used directly so number format strings are preserved
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
        logger.info(
            f"Found {len(workbook.sheetnames)} sheets: {', '.join(workbook.sheetnames)}"
        )
    except FileNotFoundError:
        logger.warning(f"Error: File not found: {excel_path}")
        return None
    except Exception as e:
        logger.warning(f"Error loading Excel file: {e}")
        return None

    sheets_encoding = {}
    compression_metrics = {"sheets": {}}
    overall_orig = overall_anchor = overall_index = overall_format = overall_final = 0

    for sheet_name in workbook.sheetnames:
        logger.info(f"\\nProcessing sheet: {sheet_name}")
        sheet = workbook[sheet_name]

        if sheet.max_row <= 1 and sheet.max_column <= 1:
            logger.info(f"Sheet '{sheet_name}' appears to be empty. Skipping.")
            continue

        logger.info(
            f"Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns"
        )
        # print memory usage
        logger.info(f"Estimated memory usage: {sys.getsizeof(sheet)} bytes")

        # --- gather original tokens before any compression ---
        original_cells = {}
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=r, column=c).value
                if cell_value is not None:
                    original_cells[f"{get_column_letter(c)}{r}"] = str(cell_value)
        original_tokens = len(json.dumps(original_cells, ensure_ascii=False))

        row_anchors, col_anchors = find_structural_anchors(sheet, k)
        logger.info(
            f"Found {len(row_anchors)} row anchors and {len(col_anchors)} column anchors"
        )

        kept_rows, kept_cols = extract_cells_near_anchors(sheet, row_anchors, col_anchors, 0)

        # Compress homogeneous regions before indexing
        kept_rows, kept_cols = compress_homogeneous_regions(sheet, kept_rows, kept_cols)
        logger.info(f"After compression: {len(kept_rows)} rows and {len(kept_cols)} columns kept")

        anchor_cells = {}
        for r in kept_rows:
            for c in kept_cols:
                cell_value = sheet.cell(row=r, column=c).value
                if cell_value is not None:
                    anchor_cells[f"{get_column_letter(c)}{r}"] = str(cell_value)
        anchor_tokens = len(json.dumps(anchor_cells, ensure_ascii=False))

        inverted_index, format_map = create_inverted_index(sheet, kept_rows, kept_cols)
        logger.info(
            f"Created inverted index with {len(inverted_index)} unique values"
        )

        merged_index = create_inverted_index_translation(inverted_index)
        logger.info(
            f"Merged values into {len(merged_index)} range groups"
        )
        index_tokens = len(json.dumps(merged_index, ensure_ascii=False))

        # Create a map from a semantic key to cell references for aggregation
        type_nfs_map = defaultdict(list)
        for _, cells in format_map.items():
            for cell_ref in cells:
                try:
                    cell = sheet[cell_ref]
                except Exception:
                    continue
                nfs = get_number_format_string(cell)
                sem_type = detect_semantic_type(cell)
                key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
                type_nfs_map[key].append(cell_ref)

        aggregated_formats = aggregate_regions_dfs(sheet, type_nfs_map)
        logger.info(
            f"Aggregated {len(aggregated_formats)} format regions"
        )
        format_tokens = len(json.dumps(aggregated_formats, ensure_ascii=False))

        numeric_map = {
            fmt: cells
            for fmt, cells in type_nfs_map.items()
            if json.loads(fmt).get("type") in ["numeric", "integer", "float"]
        }
        numeric_ranges = aggregate_regions_dfs(sheet, numeric_map)
        logger.info(f"Clustered {len(numeric_ranges)} numeric format ranges")

        sheet_encoding = {
            "structural_anchors": {
                "rows": row_anchors,
                "columns": [get_column_letter(c) for c in col_anchors]
            },
            "cells": merged_index,
            "formats": aggregated_formats,
            "numeric_ranges": numeric_ranges
        }

        final_tokens = len(json.dumps(sheet_encoding, ensure_ascii=False))

        ratio_anchor = calculate_compression_ratio(original_tokens, anchor_tokens)
        ratio_index = calculate_compression_ratio(original_tokens, index_tokens)
        ratio_format = calculate_compression_ratio(original_tokens, format_tokens)
        ratio_final = calculate_compression_ratio(original_tokens, final_tokens)

        compression_metrics["sheets"][sheet_name] = {
            "original_tokens": original_tokens,
            "after_anchor_tokens": anchor_tokens,
            "after_inverted_index_tokens": index_tokens,
            "after_format_tokens": format_tokens,
            "final_tokens": final_tokens,
            "anchor_ratio": ratio_anchor,
            "inverted_index_ratio": ratio_index,
            "format_ratio": ratio_format,
            "overall_ratio": ratio_final,
        }

        logger.info(
            f"{sheet_name} compression - Anchors: {ratio_anchor:.2f}x, "
            f"Index: {ratio_index:.2f}x, Formats: {ratio_format:.2f}x, "
            f"Overall: {ratio_final:.2f}x"
        )

        sheets_encoding[sheet_name] = sheet_encoding

        overall_orig += original_tokens
        overall_anchor += anchor_tokens
        overall_index += index_tokens
        overall_format += format_tokens
        overall_final += final_tokens

    compression_metrics["overall"] = {
        "original_tokens": overall_orig,
        "after_anchor_tokens": overall_anchor,
        "after_inverted_index_tokens": overall_index,
        "after_format_tokens": overall_format,
        "final_tokens": overall_final,
        "anchor_ratio": calculate_compression_ratio(overall_orig, overall_anchor),
        "inverted_index_ratio": calculate_compression_ratio(overall_orig, overall_index),
        "format_ratio": calculate_compression_ratio(overall_orig, overall_format),
        "overall_ratio": calculate_compression_ratio(overall_orig, overall_final),
    }

    logger.info(
        f"Overall compression: {compression_metrics['overall']['overall_ratio']:.2f}x"
    )

    full_encoding = {
        "file_name": os.path.basename(excel_path),
        "sheets": sheets_encoding,
        "compression_metrics": compression_metrics,
    }

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(full_encoding, f, indent=2, ensure_ascii=False)
        logger.info(f"Saved SpreadsheetLLM encoding to {output_path}")

    return full_encoding


def get_cell_style_key(cell):
    """Creates a hashable key representing a cell's style for comparison."""
    if not cell:
        return "no_cell"

    font = cell.font
    border = cell.border
    fill = cell.fill
    alignment = cell.alignment

    # Create a tuple of style attributes. Tuples are hashable.
    style_tuple = (
        (font.bold, font.italic, font.underline, font.sz, str(font.color.rgb if font.color else None)),
        (border.left.style, border.right.style, border.top.style, border.bottom.style),
        (fill.patternType, str(fill.fgColor.rgb if fill.fgColor else None)),
        (alignment.horizontal, alignment.vertical, alignment.wrap_text)
    )
    return style_tuple


def is_header_row(sheet, row_idx):
    """More robust heuristics to detect header rows, as per Appendix C."""
    num_populated = 0
    num_bold = 0
    num_all_caps = 0
    num_strings = 0
    num_centered = 0

    for c in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_idx, column=c)
        if cell.value is None or str(cell.value).strip() == "":
            continue

        num_populated += 1
        if cell.font and cell.font.bold:
            num_bold += 1
        if cell.alignment and cell.alignment.horizontal == 'center':
            num_centered += 1

        if isinstance(cell.value, str):
            num_strings += 1
            if cell.value.isupper() and len(cell.value) > 1:
                num_all_caps += 1

    if num_populated == 0:
        return False

    # A high proportion of bolded, centered, or all-caps text cells are strong indicators.
    if num_bold / num_populated > 0.6:
        return True
    if num_centered / num_populated > 0.6:
        return True
    if num_strings > 0 and num_all_caps / num_strings > 0.6:
        return True

    return False


def find_boundary_candidates(sheet):
    """
    Identify row/column boundary candidates using enhanced heterogeneity heuristics
    from Appendix C, including cell value, merged status, and style.
    """
    row_profiles = []
    for r in range(1, sheet.max_row + 1):
        profile = []
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            is_merged = any(cell.coordinate in r_ for r_ in sheet.merged_cells.ranges)
            style_key = get_cell_style_key(cell)
            profile.append((cell.value, is_merged, style_key))
        row_profiles.append(profile)

    col_profiles = []
    for c in range(1, sheet.max_column + 1):
        profile = []
        for r in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=c)
            is_merged = any(cell.coordinate in r_ for r_ in sheet.merged_cells.ranges)
            style_key = get_cell_style_key(cell)
            profile.append((cell.value, is_merged, style_key))
        col_profiles.append(profile)

    row_candidates = set()
    for r in range(1, len(row_profiles)):
        if row_profiles[r] != row_profiles[r - 1]:
            # Add both sides of the boundary
            row_candidates.add(r)
            row_candidates.add(r + 1)

    col_candidates = set()
    for c in range(1, len(col_profiles)):
        if col_profiles[c] != col_profiles[c - 1]:
            col_candidates.add(c)
            col_candidates.add(c + 1)

    # Filter out candidates that are part of a detected header region
    header_rows = {idx for idx in range(1, sheet.max_row + 1) if is_header_row(sheet, idx)}
    row_candidates = {r for r in row_candidates if r not in header_rows}

    # Step 2: Compose candidate boundaries
    candidates = []
    if row_candidates and col_candidates:
        rows = sorted(list(row_candidates))
        cols = sorted(list(col_candidates))
        for i in range(len(rows)):
            for j in range(i + 1, len(rows)):
                for k in range(len(cols)):
                    for l in range(k + 1, len(cols)):
                        candidates.append((rows[i], cols[k], rows[j], cols[l]))

    # Step 3: Filter unreasonable candidates
    candidates = filter_unreasonable_candidates(sheet, candidates)

    # Step 4: Filter overlapping candidates
    candidates = filter_overlapping_candidates(sheet, candidates)

    # Step 5: Derive anchors from final candidates
    final_row_anchors = set()
    final_col_anchors = set()
    for r1, c1, r2, c2 in candidates:
        final_row_anchors.add(r1)
        final_row_anchors.add(r2)
        final_col_anchors.add(c1)
        final_col_anchors.add(c2)

    return sorted(list(final_row_anchors)), sorted(list(final_col_anchors))


def filter_unreasonable_candidates(sheet, candidates):
    """Filter out candidates based on size, sparsity, and header presence."""
    filtered = []
    for r1, c1, r2, c2 in candidates:
        # Size filter
        if (r2 - r1 < 1) or (c2 - c1 < 1): continue # Must have at least 2 rows/cols

        # Sparsity filter
        num_cells = (r2 - r1 + 1) * (c2 - c1 + 1)
        populated_cells = 0
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if sheet.cell(row=r, column=c).value is not None:
                    populated_cells += 1

        if populated_cells / num_cells < 0.1: # At least 10% populated
            continue

        # Header presence filter (simple version)
        has_header = any(is_header_row(sheet, r) for r in range(r1, r2 + 1))
        if not has_header:
            continue

        filtered.append((r1, c1, r2, c2))

    return filtered


def calculate_iou(box1, box2):
    """Calculate Intersection over Union (IoU) for two bounding boxes."""
    r1_1, c1_1, r2_1, c2_1 = box1
    r1_2, c1_2, r2_2, c2_2 = box2

    inter_r1 = max(r1_1, r1_2)
    inter_c1 = max(c1_1, c1_2)
    inter_r2 = min(r2_1, r2_2)
    inter_c2 = min(c2_1, c2_2)

    inter_area = max(0, inter_r2 - inter_r1 + 1) * max(0, inter_c2 - inter_c1 + 1)

    area1 = (r2_1 - r1_1 + 1) * (c2_1 - c1_1 + 1)
    area2 = (r2_2 - r1_2 + 1) * (c2_2 - c1_2 + 1)

    union_area = area1 + area2 - inter_area

    return inter_area / union_area if union_area > 0 else 0


def filter_overlapping_candidates(sheet, candidates):
    """Filter overlapping candidates using heuristics from Appendix C."""
    if not candidates:
        return []

    # Score candidates (higher is better)
    scores = []
    for r1, c1, r2, c2 in candidates:
        score = 0
        # Header score
        for r in range(r1, min(r1 + 3, r2 + 1)): # Check top 3 rows for header
            if is_header_row(sheet, r):
                score += 10
        # Area score
        score += (r2 - r1 + 1) * (c2 - c1 + 1)
        scores.append(score)

    # Non-maximum suppression based on IoU and scores
    indices = list(range(len(candidates)))
    indices.sort(key=lambda i: scores[i], reverse=True)

    keep = []
    while indices:
        current_idx = indices.pop(0)
        keep.append(current_idx)

        remaining_indices = []
        for idx in indices:
            iou = calculate_iou(candidates[current_idx], candidates[idx])
            # If high overlap, discard the one with the lower score (which is the current `idx` because of sorting)
            if iou < 0.5:
                remaining_indices.append(idx)
        indices = remaining_indices

    return [candidates[i] for i in keep]


def extract_k_neighborhood(indices, k, max_index):
    """Expand indices with a k-neighborhood within bounds."""
    expanded = set()
    for idx in indices:
        for i in range(max(1, idx - k), min(max_index + 1, idx + k + 1)):
            expanded.add(i)
    return sorted(expanded)


def find_structural_anchors(sheet, k=2):
    """Find structural anchors using boundary candidates and k-neighborhood."""
    row_candidates, col_candidates = find_boundary_candidates(sheet)
    row_anchors = extract_k_neighborhood(row_candidates, k, sheet.max_row)
    col_anchors = extract_k_neighborhood(col_candidates, k, sheet.max_column)
    return row_anchors, col_anchors


def get_cell_format_key(cell):
    """Helper function to create a consistent format key for a cell."""
    format_info = {}
    try:
        # 1. Font Styles
        font = cell.font
        format_info["font"] = {
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "name": font.name,
            "size": font.sz,
            "color": str(font.color.rgb) if font.color and font.color.rgb else None,  # Get RGB color
        }

        # 2. Alignment
        alignment = cell.alignment
        format_info["alignment"] = {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
        }

        # 3. Borders
        border = cell.border
        format_info["border"] = {
            side: {
                "style": getattr(border, side).style,
                "color": (
                    str(getattr(border, side).color.rgb)
                    if getattr(border, side).color and getattr(border, side).color.rgb
                    else None
                ),
            }
            for side in ["left", "right", "top", "bottom"]
        }

        # 4. Fill (Background Color)
        fill = cell.fill
        if hasattr(fill, 'patternType') and fill.patternType == "solid":
            format_info["fill"] = {"color": str(fill.start_color.index)
                                   if fill.start_color and fill.start_color.index else None}
        else:
            format_info["fill"] = {"color": None}

        # 5. Number Format (Original, Inferred Type, Category)
        original_number_format = cell.number_format
        inferred_type = infer_cell_data_type(cell)  # Call new helper
        category = categorize_number_format(original_number_format, inferred_type)  # Call new helper

        format_info["original_number_format"] = original_number_format
        format_info["inferred_data_type"] = inferred_type
        format_info["number_format_category"] = category
        # format_info["number_format"] = cell.number_format # Keep original for now, or decide if it's redundant
    except Exception as e:
        # If there's an error extracting format, use a simplified format key
        format_info = {"error": str(e)}

    return json.dumps(format_info, sort_keys=True)


def extract_cells_near_anchors(sheet, row_anchors, col_anchors, k):
    """Extract cells within k units of any anchor."""
    rows_to_keep = set()
    cols_to_keep = set()

    for r in row_anchors:
        for i in range(max(1, r - k), min(sheet.max_row + 1, r + k + 1)):
            rows_to_keep.add(i)

    for c in col_anchors:
        for i in range(max(1, c - k), min(sheet.max_column + 1, c + k + 1)):
            cols_to_keep.add(i)

    return sorted(list(rows_to_keep)), sorted(list(cols_to_keep))


def compress_homogeneous_regions(sheet, rows, cols):
    """Remove rows and columns that are homogeneous in value and format."""
    def row_homogeneous(r):
        vals = []
        fmts = []
        for c in cols:
            cell = sheet.cell(row=r, column=c)
            vals.append(cell.value)
            fmts.append(cell.number_format)
        return len(set(vals)) <= 1 and len(set(fmts)) <= 1

    def col_homogeneous(c):
        vals = []
        fmts = []
        for r in rows:
            cell = sheet.cell(row=r, column=c)
            vals.append(cell.value)
            fmts.append(cell.number_format)
        return len(set(vals)) <= 1 and len(set(fmts)) <= 1

    filtered_rows = [r for r in rows if not row_homogeneous(r)]
    filtered_cols = [c for c in cols if not col_homogeneous(c)]
    return filtered_rows, filtered_cols


def create_inverted_index(sheet, kept_rows, kept_cols):
    """Create an inverted index, handling merged cells."""
    inverted_index = defaultdict(list)
    format_map = defaultdict(list)
    merged_ranges = sheet.merged_cells.ranges  # get all merged cell ranges

    for row in kept_rows:
        for col in kept_cols:
            cell = sheet.cell(row=row, column=col)
            cell_ref = f"{get_column_letter(col)}{row}"

            # Merged Cell Handling
            merged_value = None
            merged_range = None
            for m_range in merged_ranges:
                if cell_ref in m_range:
                    try:
                        merged_value = sheet[m_range.start_cell.coordinate].value
                        merged_range = m_range
                        break
                    except Exception:
                        pass  # Skip if there's an issue with the merged range

            # Use merged value if available, otherwise cell value
            try:
                if merged_value is not None:
                    cell_value = str(merged_value) if merged_value is not None else ""
                    inverted_index[cell_value].append(cell_ref)
                elif cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        cell_value = f"{cell.value}"
                    else:
                        cell_value = str(cell.value)
                    inverted_index[cell_value].append(cell_ref)
            except Exception as e:
                # Handle error for problematic cell values
                logger.warning(f"Error processing cell {cell_ref}: {e}")
                cell_value = "ERROR_VALUE"
                inverted_index[cell_value].append(cell_ref)

            # Format Handling
            try:
                format_info = {}
                # 1. Font Styles
                font = cell.font
                format_info["font"] = {
                    "bold": font.bold,
                    "italic": font.italic,
                    "underline": font.underline,
                    "name": font.name,
                    "size": font.sz,
                    "color": str(font.color.rgb) if font.color and font.color.rgb else None,
                }

                # 2. Alignment
                alignment = cell.alignment
                format_info["alignment"] = {
                    "horizontal": alignment.horizontal,
                    "vertical": alignment.vertical,
                }

                # 3. Borders
                border = cell.border
                format_info["border"] = {
                    side: {
                        "style": getattr(border, side).style,
                        "color": (
                            str(getattr(border, side).color.rgb)
                            if getattr(border, side).color and getattr(border, side).color.rgb
                            else None
                        ),
                    }
                    for side in ["left", "right", "top", "bottom"]
                }

                # 4. Fill (Background Color)
                fill = cell.fill
                if hasattr(fill, 'patternType') and fill.patternType == "solid":
                    format_info["fill"] = {"color": str(fill.start_color.index)
                                           if fill.start_color and fill.start_color.index else None}
                else:
                    format_info["fill"] = {"color": None}

                # 5. Number Format (Original, Inferred Type, Category)
                original_number_format = cell.number_format
                inferred_type = infer_cell_data_type(cell)
                category = categorize_number_format(original_number_format, cell)

                format_info["original_number_format"] = original_number_format
                format_info["inferred_data_type"] = inferred_type
                format_info["number_format_category"] = category

                # Store the format (handle merged ranges specially in format key)
                if merged_range is not None:
                    format_info["merged"] = True
                    format_info["merged_range"] = str(merged_range)
                else:
                    format_info["merged"] = False

                format_key = json.dumps(format_info, sort_keys=True)
                format_map[format_key].append(cell_ref)
            except Exception as e:
                # Handle error for problematic cell formats
                logger.warning(f"Error processing format for cell {cell_ref}: {e}")

    return dict(inverted_index), dict(format_map)


def create_inverted_index_translation(inverted_index):
    """Merge cell references for identical values into ranges.

    Args:
        inverted_index (dict): Mapping of values to lists of cell references.

    Returns:
        dict: Mapping of values to merged cell ranges.
    """

    def _merge_refs(refs):
        coords = []
        for ref in sorted(set(refs)):
            try:
                col_letter, row = split_cell_ref(ref)
                col = openpyxl.utils.cell.column_index_from_string(col_letter)
                coords.append((row, col))
            except Exception:
                continue

        cell_set = set(coords)
        processed = set()
        ranges = []

        for row, col in sorted(coords):
            if (row, col) in processed:
                continue

            width = 1
            while (row, col + width) in cell_set and (row, col + width) not in processed:
                width += 1

            height = 1
            expanding = True
            while expanding:
                next_row = row + height
                for w in range(width):
                    if (next_row, col + w) not in cell_set or (next_row, col + w) in processed:
                        expanding = False
                        break
                if expanding:
                    height += 1

            end_col = col + width - 1
            end_row = row + height - 1
            start_ref = f"{get_column_letter(col)}{row}"
            end_ref = f"{get_column_letter(end_col)}{end_row}"

            if width == 1 and height == 1:
                ranges.append(start_ref)
            else:
                ranges.append(f"{start_ref}:{end_ref}")

            for r in range(row, row + height):
                for c in range(col, col + width):
                    processed.add((r, c))

        return ranges

    merged_index = {}
    for value, refs in inverted_index.items():
        if value is None or str(value).strip() == "":
            continue
        merged_index[value] = _merge_refs(refs)

    return merged_index


def aggregate_formats(sheet, format_map):
    """Aggregate cells with the same inferred type and number format string."""
    aggregated_formats = defaultdict(list)
    processed_cells = set()

    type_nfs_map = defaultdict(list)
    for _, cells in format_map.items():
        for cell_ref in cells:
            try:
                cell = sheet[cell_ref]
            except Exception:
                continue
            nfs = get_number_format_string(cell)
            sem_type = detect_semantic_type(cell)
            key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
            type_nfs_map[key].append(cell_ref)

    for key, cells in type_nfs_map.items():
        cells_set = set(cells)
        for start_cell in cells:
            if start_cell in processed_cells:
                continue

            try:
                start_col_letter, start_row = split_cell_ref(start_cell)
                start_col = openpyxl.utils.cell.column_index_from_string(start_col_letter)
            except Exception:
                continue

            best_width = 1
            best_height = 1
            best_area = 1
            best_end_cell = start_cell

            max_width = min(20, sheet.max_column - start_col + 1)
            max_height = min(20, sheet.max_row - start_row + 1)

            for width in range(1, max_width + 1):
                for height in range(1, max_height + 1):
                    valid_rectangle = True
                    for r in range(start_row, start_row + height):
                        for c in range(start_col, start_col + width):
                            cell_ref = f"{get_column_letter(c)}{r}"
                            if cell_ref not in cells_set or cell_ref in processed_cells:
                                valid_rectangle = False
                                break
                        if not valid_rectangle:
                            break

                    if valid_rectangle:
                        area = width * height
                        if area > best_area:
                            best_width = width
                            best_height = height
                            best_area = area
                            best_end_cell = f"{get_column_letter(start_col + width - 1)}{start_row + height - 1}"

            region = start_cell if best_width == 1 and best_height == 1 else f"{start_cell}:{best_end_cell}"
            aggregated_formats[key].append(region)
            for r in range(start_row, start_row + best_height):
                for c in range(start_col, start_col + best_width):
                    processed_cells.add(f"{get_column_letter(c)}{r}")

    return dict(aggregated_formats)


def cluster_numeric_ranges(sheet, format_map):
    """Aggregate numeric cells with identical formatting into ranges."""
    numeric_map = {
        fmt: cells
        for fmt, cells in format_map.items()
        if json.loads(fmt).get("inferred_data_type") == "numeric"
    }

    if not numeric_map:
        # Fall back to scanning the entire sheet when no anchors were
        # retained and format_map is empty. This ensures numeric ranges are
        # still detected for simple numeric sheets.
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if infer_cell_data_type(cell) == "numeric":
                    fmt_key = json.dumps(
                        {
                            "type": detect_semantic_type(cell),
                            "nfs": get_number_format_string(cell),
                        },
                        sort_keys=True,
                    )
                    numeric_map.setdefault(fmt_key, []).append(
                        f"{get_column_letter(c)}{r}"
                    )

    return aggregate_formats(sheet, numeric_map)


def get_column_index(col_letter):
    """Convert column letter to index (A => 1, AA => 27)."""
    return openpyxl.utils.cell.column_index_from_string(col_letter)


def split_cell_ref(cell_ref):
    """Split cell reference (e.g., 'A1') into column letter and row number."""
    col_str = ''.join(filter(str.isalpha, cell_ref))
    row_str = ''.join(filter(str.isdigit, cell_ref))

    # convert row to integer
    return col_str, int(row_str)


def main():
    """Console script entry point for SpreadsheetLLM encoder."""
    logging.basicConfig(level=logging.INFO)
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert Excel files to SpreadsheetLLM format"
    )
    parser.add_argument("excel_file", help="Path to the Excel file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output JSON file path (default: same as input with .json extension)",
    )
    parser.add_argument(
        "--k",
        type=int,
        default=2,
        help="Neighborhood distance parameter (default: 2)",
    )
    parser.add_argument(
        "--vanilla",
        action="store_true",
        help="Produce vanilla markdown-like encoding instead of compressed JSON.",
    )

    args = parser.parse_args()

    if not args.output:
        if args.vanilla:
            args.output = os.path.splitext(args.excel_file)[0] + "_vanilla.txt"
        else:
            args.output = os.path.splitext(args.excel_file)[0] + "_spreadsheetllm.json"

    spreadsheet_llm_encode(args.excel_file, args.output, args.k, args.vanilla)


def vanilla_encode(excel_path, output_path=None):
    """
    Produces a simple vanilla markdown-like encoding of a spreadsheet.
    """
    logger.info(f"Producing vanilla encoding for {excel_path}")
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        logger.error(f"Error loading Excel file for vanilla encoding: {e}")
        return None

    vanilla_content = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_str = []
        for r in range(1, sheet.max_row + 1):
            row_str = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                cell_ref = f"{get_column_letter(c)}{r}"
                cell_val = str(cell.value) if cell.value is not None else ""
                row_str.append(f"{cell_ref},{cell_val}")
            sheet_str.append("|".join(row_str))
        vanilla_content[sheet_name] = "\n".join(sheet_str)

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            # For simplicity, we'll save the first sheet's content if there are multiple
            first_sheet_name = next(iter(vanilla_content))
            f.write(vanilla_content[first_sheet_name])
        logger.info(f"Saved vanilla encoding to {output_path}")

    return vanilla_content


if __name__ == "__main__":
    main()

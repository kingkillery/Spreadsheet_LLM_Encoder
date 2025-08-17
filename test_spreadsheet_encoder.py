import unittest
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from Spreadsheet_LLM_Encoder import spreadsheet_llm_encode, find_boundary_candidates, aggregate_regions_dfs, vanilla_encode

class TestSpreadsheetEncoder(unittest.TestCase):

    def setUp(self):
        """Set up a test workbook."""
        self.test_file = "test_workbook.xlsx"
        wb = openpyxl.Workbook()

        # Sheet 1: For boundary and aggregation tests
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1['A1'] = "Header 1"
        ws1['A1'].font = Font(bold=True)
        ws1['B1'] = "Header 2"
        ws1['B1'].font = Font(bold=True)
        ws1['A2'] = 100
        ws1['B2'] = 200
        ws1['A3'] = 150
        ws1['B3'] = 250

        # Add a separate region with a different format
        ws1['D4'] = "Data"
        ws1['D4'].fill = PatternFill("solid", fgColor="FFFF00")
        ws1['E4'] = "More Data"
        ws1['E4'].fill = PatternFill("solid", fgColor="FFFF00")

        # Sheet 2: For vanilla encoding test
        ws2 = wb.create_sheet("Sheet2")
        ws2['A1'] = "Hello"
        ws2['B1'] = "World"

        wb.save(self.test_file)

    def tearDown(self):
        """Remove the test workbook."""
        os.remove(self.test_file)

    def test_vanilla_encode(self):
        result = vanilla_encode(self.test_file)
        self.assertIn("Sheet1", result)
        self.assertIn("Sheet2", result)
        self.assertTrue(result["Sheet1"].startswith("A1,Header 1|B1,Header 2"))
        self.assertTrue(result["Sheet2"].startswith("A1,Hello|B1,World"))

    def test_find_boundary_candidates_advanced(self):
        wb = openpyxl.load_workbook(self.test_file)
        sheet = wb["Sheet1"]
        rows, cols = find_boundary_candidates(sheet)
        # This is a basic check; the full heuristics are complex.
        # We expect a boundary between the two groups of cells.
        self.assertIn(4, rows) # Boundary between row 3 and 4
        self.assertIn(3, cols) # Boundary between col B and C

    def test_aggregate_regions_dfs(self):
        wb = openpyxl.load_workbook(self.test_file)
        sheet = wb["Sheet1"]

        format_map = {
            json.dumps({"type": "integer", "nfs": "General"}, sort_keys=True): ["A2", "B2", "A3", "B3"],
            json.dumps({"type": "text", "nfs": "General"}, sort_keys=True): ["D4", "E4"]
        }

        aggregated = aggregate_regions_dfs(sheet, format_map)

        # Check that the regions were aggregated correctly
        key1 = json.dumps({"type": "integer", "nfs": "General"}, sort_keys=True)
        key2 = json.dumps({"type": "text", "nfs": "General"}, sort_keys=True)
        self.assertIn("A2:B3", aggregated[key1])
        self.assertIn("D4:E4", aggregated[key2])

    def test_spreadsheet_llm_encode_runs(self):
        result = spreadsheet_llm_encode(self.test_file)
        self.assertIsNotNone(result)
        self.assertIn("Sheet1", result["sheets"])

if __name__ == '__main__':
    unittest.main()

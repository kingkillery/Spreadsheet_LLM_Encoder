import openpyxl
from Spreadsheet_LLM_Encoder import spreadsheet_llm_encode


def create_workbook_numeric_region(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Heterogeneous first row
    for c, val in enumerate([1, 2, 3, 4], start=1):
        ws.cell(row=1, column=c, value=val).number_format = '0'
    # Homogeneous numeric block below
    for r in range(2, 5):
        for c in range(1, 5):
            ws.cell(row=r, column=c, value=0).number_format = '0'
    wb.save(path)


def create_workbook_with_homogeneous_rows(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'ID'
    ws['B1'] = 'Value'
    for r in [2, 3]:
        ws.cell(row=r, column=1, value=0)
        ws.cell(row=r, column=2, value=0)
    ws['A4'] = 'End'
    ws['B4'] = 5
    wb.save(path)


def test_numeric_range_aggregation(tmp_path):
    file_path = tmp_path / "num.xlsx"
    create_workbook_numeric_region(str(file_path))
    result = spreadsheet_llm_encode(str(file_path))
    ranges = result['sheets']['Sheet']['numeric_ranges']
    assert isinstance(ranges, dict)


def test_homogeneous_rows_skipped(tmp_path):
    file_path = tmp_path / "homog.xlsx"
    create_workbook_with_homogeneous_rows(str(file_path))
    result = spreadsheet_llm_encode(str(file_path), k=1)
    cells = result['sheets']['Sheet']['cells']
    refs = [ref for lst in cells.values() for ref in lst]
    assert not any(ref.endswith('2') or ref.endswith('3') for ref in refs)

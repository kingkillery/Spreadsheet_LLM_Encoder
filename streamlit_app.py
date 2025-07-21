import streamlit as st
import pandas as pd
import json
import base64
import logging
from Spreadsheet_LLM_Encoder import spreadsheet_llm_encode
import openpyxl
import tempfile
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string # For parsing cell/range refs
import re # For parsing number format strings
from collections import Counter # For compression insights (value frequency) - not strictly needed for current value_counts but good for future

def get_download_link(json_data, filename="spreadsheet_data.json"):
    """
    Generates an HTML download link for JSON data.

    Args:
        json_data (dict or list): The JSON data to be downloaded.
        filename (str, optional): The desired filename for the downloaded file.
                                  Defaults to "spreadsheet_data.json".

    Returns:
        str: An HTML string representing the download link.
    """
    json_str = json.dumps(json_data, indent=2)
    b64 = base64.b64encode(json_str.encode()).decode()
    href = f'<a href="data:file/json;base64,{b64}" download="{filename}">Download JSON File</a>'
    return href

def detect_tables_in_sheet(sheet):
    """
    Detects tables in a given Openpyxl worksheet object using heuristics.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to analyze.

    Heuristics:
        - Header Identification: A row is a potential header if a significant
          proportion of its populated cells are bold, have a bottom border, or
          if all populated cells are strings and a significant proportion are all uppercase.
        - Table Body: Subsequent rows are part of the table if they are not empty
          and maintain a similar number of populated cells and column span as the header.
        - Table End: A table ends when an empty row is encountered or the data
          pattern (column alignment, cell density) breaks significantly.

    Returns:
        list: A list of dictionaries, where each dictionary represents a detected table.
              Each dictionary contains "full_range", "header_range", "data_range",
              and "detection_method" keys. Returns an empty list if no tables are found.
    """
    if not sheet:
        return []

    tables = []

    try:
        min_row, min_col, max_row, max_col = sheet.min_row, sheet.min_column, sheet.max_row, sheet.max_column
    except Exception:
        return []

    if max_row == 0 or max_col == 0 :
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1,1).value is None:
             return []
        try:
            sheet.calculate_dimension(force=True)
            min_row, min_col, max_row, max_col = sheet.min_row, sheet.min_column, sheet.max_row, sheet.max_column
            if max_row == 0 or max_col == 0 : return []
        except:
            return []


    r_idx = min_row
    while r_idx <= max_row:
        num_populated_cells_in_row = 0
        num_bold = 0
        num_bottom_border = 0
        num_all_caps_strings = 0
        num_string_cells = 0
        current_row_min_col = max_col + 1
        current_row_max_col = min_col - 1

        for c_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                num_populated_cells_in_row += 1
                current_row_min_col = min(current_row_min_col, c_idx)
                current_row_max_col = max(current_row_max_col, c_idx)

                if cell.font and cell.font.bold:
                    num_bold += 1
                if cell.border and cell.border.bottom and cell.border.bottom.style is not None and cell.border.bottom.style != 'none':
                    num_bottom_border += 1
                if isinstance(cell.value, str):
                    num_string_cells += 1
                    if cell.value and cell.value.isupper():
                        num_all_caps_strings += 1

        is_header_candidate = False
        if num_populated_cells_in_row > 1:
            if (num_bold / num_populated_cells_in_row > 0.5) or \
               (num_bottom_border / num_populated_cells_in_row > 0.5) or \
               (num_string_cells == num_populated_cells_in_row and \
                num_string_cells > 0 and \
                (num_all_caps_strings / num_string_cells > 0.5)):
                is_header_candidate = True
        elif num_populated_cells_in_row == 1:
             if num_bold == 1 or num_bottom_border == 1 or (num_string_cells == 1 and num_all_caps_strings ==1):
                 is_header_candidate = True

        if is_header_candidate:
            header_row_idx = r_idx
            header_actual_min_col = current_row_min_col
            header_actual_max_col = current_row_max_col
            table_min_col_overall = header_actual_min_col
            table_max_col_overall = header_actual_max_col
            data_end_row_idx = header_row_idx

            for data_r_idx in range(header_row_idx + 1, max_row + 1):
                data_row_populated_cells = 0
                data_row_min_col_actual = max_col + 1
                data_row_max_col_actual = min_col - 1

                for c_idx_data in range(min_col, max_col + 1):
                    cell_data = sheet.cell(row=data_r_idx, column=c_idx_data)
                    if cell_data.value is not None and str(cell_data.value).strip() != "":
                        data_row_populated_cells +=1
                        data_row_min_col_actual = min(data_row_min_col_actual, c_idx_data)
                        data_row_max_col_actual = max(data_row_max_col_actual, c_idx_data)

                if data_row_populated_cells == 0:
                    break

                if data_row_populated_cells > 0 and \
                   (data_row_max_col_actual >= table_min_col_overall and data_row_min_col_actual <= table_max_col_overall) :
                    data_end_row_idx = data_r_idx
                    table_min_col_overall = min(table_min_col_overall, data_row_min_col_actual)
                    table_max_col_overall = max(table_max_col_overall, data_row_max_col_actual)
                else:
                    break

            if data_end_row_idx >= header_row_idx:
                header_range_str = f"{get_column_letter(table_min_col_overall)}{header_row_idx}:{get_column_letter(table_max_col_overall)}{header_row_idx}"
                data_range_str = None
                if data_end_row_idx > header_row_idx:
                    data_range_str = f"{get_column_letter(table_min_col_overall)}{header_row_idx + 1}:{get_column_letter(table_max_col_overall)}{data_end_row_idx}"
                full_range_str = f"{get_column_letter(table_min_col_overall)}{header_row_idx}:{get_column_letter(table_max_col_overall)}{data_end_row_idx}"

                tables.append({
                    "full_range": full_range_str,
                    "header_range": header_range_str,
                    "data_range": data_range_str,
                    "detection_method": "improved_heuristic_v1"
                })
                r_idx = data_end_row_idx + 1
                continue
        r_idx += 1
    return tables

def extract_chart_info_from_sheet(sheet):
    """
    Extracts basic information about charts in a given Openpyxl worksheet.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to analyze.

    Returns:
        list: A list of dictionaries, where each dictionary contains information
              about a chart (title, type, anchor_cell). Returns an empty list
              if no charts are found or if chart information cannot be extracted.
    """
    chart_info_list = []
    if not hasattr(sheet, '_charts') or not sheet._charts:
        return chart_info_list

    for chart in sheet._charts:
        title_str = None
        if chart.title:
            if hasattr(chart.title, 'tx') and hasattr(chart.title.tx, 'rich') and \
               hasattr(chart.title.tx.rich, 'p') and len(chart.title.tx.rich.p) > 0 and \
               hasattr(chart.title.tx.rich.p[0], 'r') and len(chart.title.tx.rich.p[0].r) > 0 and \
               hasattr(chart.title.tx.rich.p[0].r[0], 't'):
                title_str = chart.title.tx.rich.p[0].r[0].t
            elif isinstance(chart.title, str):
                title_str = chart.title

        chart_type_str = chart.type if chart.type else None
        anchor_cell_str = None

        try:
            if hasattr(chart, 'anchor'):
                anchor = chart.anchor
                if hasattr(anchor, 'twoCellAnchor'):
                    _from = anchor.twoCellAnchor._from
                    anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
                elif hasattr(anchor, 'oneCellAnchor'):
                    _from = getattr(anchor.oneCellAnchor, '_from', None)
                    if _from and hasattr(_from, 'col') and hasattr(_from, 'row'):
                         anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
                    elif hasattr(anchor.oneCellAnchor, 'cell') and hasattr(anchor.oneCellAnchor.cell, 'coordinate'):
                        anchor_cell_str = anchor.oneCellAnchor.cell.coordinate
                elif hasattr(anchor, '_from') and hasattr(anchor._from, 'col') and hasattr(anchor._from, 'row'):
                    _from = anchor._from
                    anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
        except Exception:
            pass

        chart_info_list.append({
            "title": title_str,
            "type": chart_type_str,
            "anchor": anchor_cell_str
        })
    return chart_info_list

def parse_number_format_string(format_str):
    """
    Parses an Excel number format string to extract type and details.

    Args:
        format_str (str): The number format string (e.g., "[$USD]#,##0.00", "0.00%").

    Returns:
        dict: A dictionary containing the parsed information, including "type",
              "original" format string, and other relevant details like "symbol",
              "decimals", "format_tokens", or "thousands_separator".
    """
    if format_str is None:
        return {"type": "none", "original": None}
    if not isinstance(format_str, str):
        return {"type": "invalid", "original": str(format_str)}

    original_format_str = format_str
    format_str_lower = format_str.lower()

    if format_str_lower == "general":
        return {"type": "general", "original": original_format_str}
    if format_str_lower == "@" or format_str_lower == "text":
        return {"type": "text", "original": original_format_str}

    if "%" in format_str:
        decimals = 0
        m = re.search(r'\.([0#]+)(?=[^%]*%)', format_str)
        if m:
            decimals = len(m.group(1))
        return {"type": "percentage", "decimals": decimals, "original": original_format_str}

    currency_match = re.search(r'(?<!\[[^\]]*)([$€£¥])', format_str)
    if currency_match:
        symbol_map = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "JPY"}
        found_symbol_char = currency_match.group(1)
        currency_code = symbol_map.get(found_symbol_char, found_symbol_char)

        decimals = 0
        m = re.search(r'\.([0#]+)', format_str)
        if m:
            decimals = len(m.group(1))
        return {"type": "currency", "symbol": currency_code, "decimals": decimals, "original": original_format_str}

    has_date_chars = any(c in format_str_lower for c in ['y', 'd']) or "mmm" in format_str_lower
    has_time_chars = any(c in format_str_lower for c in ['h', 's', 'am/pm', 'ss'])

    is_date = False
    is_time = False

    if has_date_chars:
        is_date = True
    if has_time_chars:
        is_time = True

    if 'm' in format_str_lower and not is_date and not is_time:
        if re.fullmatch(r'm{1,5}', format_str_lower):
            is_date = True
        elif ('h' in format_str_lower or 's' in format_str_lower) and 'm' in format_str_lower:
             is_time = True

    if is_date and is_time:
        return {"type": "datetime", "format_tokens": original_format_str, "original": original_format_str}
    if is_date:
        return {"type": "date", "format_tokens": original_format_str, "original": original_format_str}
    if is_time:
        return {"type": "time", "format_tokens": original_format_str, "original": original_format_str}

    if "e+" in format_str_lower or "e-" in format_str_lower:
        decimals = 0
        m = re.search(r'\.([0#]+)(?=E)', format_str, re.IGNORECASE)
        if m:
            decimals = len(m.group(1))
        return {"type": "scientific", "decimals": decimals, "original": original_format_str}

    if "/" in format_str and "?" in format_str:
        return {"type": "fraction", "original": original_format_str}

    if re.fullmatch(r'[#0,]+(\.[#0]+)?', format_str) or re.fullmatch(r'[#0,]+', format_str):
        decimals = 0
        if "." in format_str:
            m = re.search(r'\.([0#]+)', format_str)
            if m:
                decimals = len(m.group(1))
        has_comma = "," in format_str.split(".")[0]
        return {"type": "number", "decimals": decimals, "thousands_separator": has_comma, "original": original_format_str}

    return {"type": "unknown", "original": original_format_str}

def extract_sheet_metadata(sheet):
    """
    Extracts sheet-level metadata (visibility, protection status, tab color).

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to analyze.

    Returns:
        dict: A dictionary containing sheet metadata.
    """
    visibility = sheet.sheet_state if hasattr(sheet, 'sheet_state') else 'visible'
    is_protected = sheet.protection.sheet if hasattr(sheet, 'protection') and hasattr(sheet.protection, 'sheet') else False

    tab_color_str = None
    if hasattr(sheet, 'sheet_properties') and hasattr(sheet.sheet_properties, 'tabColor') and sheet.sheet_properties.tabColor:
        tab_color_obj = sheet.sheet_properties.tabColor
        if hasattr(tab_color_obj, 'rgb') and tab_color_obj.rgb:
            tab_color_str = str(tab_color_obj.rgb)
        elif hasattr(tab_color_obj, 'indexed') and tab_color_obj.indexed > 0 :
            tab_color_str = f"indexed_{tab_color_obj.indexed}"
        elif hasattr(tab_color_obj, 'theme') and hasattr(tab_color_obj, 'tint'):
             tab_color_str = f"theme_{tab_color_obj.theme}_tint_{tab_color_obj.tint}"

    return {
        "visibility": visibility,
        "is_protected": is_protected,
        "tab_color": tab_color_str
    }

def analyze_sheet_for_compression_insights(sheet_json_data):
    """
    Analyzes sheet JSON data for potential compression insights.

    Args:
        sheet_json_data (dict): The JSON data for a single sheet from the encoder's output.

    Returns:
        dict: A dictionary containing analysis insights regarding format redundancy,
              value frequency, anchor usage, and cell counts.
    """
    insights = {}

    if "format_regions" in sheet_json_data and isinstance(sheet_json_data["format_regions"], dict):
        format_keys_json_strings = list(sheet_json_data["format_regions"].keys())
        num_unique_formats_overall = len(format_keys_json_strings)
        potential_redundancy_groups = []
        base_format_groups = {}

        for fmt_key_str in format_keys_json_strings:
            try:
                fmt_details = json.loads(fmt_key_str)
                core_props_dict = {
                    "font": fmt_details.get("font"),
                    "alignment": fmt_details.get("alignment"),
                    "border": fmt_details.get("border"),
                    "fill": fmt_details.get("fill")
                }
                def make_hashable_format_core(obj):
                    if isinstance(obj, dict):
                        return tuple(sorted((k, make_hashable_format_core(v)) for k, v in obj.items()))
                    if isinstance(obj, list):
                        return tuple(make_hashable_format_core(elem) for elem in obj)
                    return obj
                base_format_key_tuple = make_hashable_format_core(core_props_dict)
                if base_format_key_tuple not in base_format_groups:
                    base_format_groups[base_format_key_tuple] = []
                base_format_groups[base_format_key_tuple].append(fmt_details.get("number_format", "General"))
            except json.JSONDecodeError:
                logger.warning(
                    f"Warning (compression_insights): Malformed format key string found: {fmt_key_str}"
                )
                continue

        for base_fmt_tuple, number_format_list in base_format_groups.items():
            if len(number_format_list) > 1:
                unique_number_formats_in_group = sorted(list(set(number_format_list)))
                if len(unique_number_formats_in_group) > 1:
                    potential_redundancy_groups.append({
                        "base_format_properties_hash": str(base_fmt_tuple),
                        "differing_number_formats": unique_number_formats_in_group,
                        "count_of_full_format_keys_in_group": len(number_format_list)
                    })
        insights["format_analysis"] = {
            "num_unique_formats_overall": num_unique_formats_overall,
            "num_base_format_groups": len(base_format_groups),
            "potential_redundancy_groups": sorted(potential_redundancy_groups, key=lambda x: x["count_of_full_format_keys_in_group"], reverse=True)
        }

    if "cells" in sheet_json_data and isinstance(sheet_json_data["cells"], dict):
        value_counts = {value: len(refs) for value, refs in sheet_json_data["cells"].items()}
        sorted_values_by_freq = sorted(value_counts.items(), key=lambda item: (item[1], str(item[0])), reverse=True)
        insights["value_frequency"] = {
            "num_unique_values": len(sorted_values_by_freq),
            "top_10_frequent_values": sorted_values_by_freq[:10]
        }
        total_cell_references = sum(len(refs) for refs in sheet_json_data["cells"].values())
        insights["total_cell_references_in_index"] = total_cell_references

    if "structural_anchors" in sheet_json_data and isinstance(sheet_json_data["structural_anchors"], dict):
        insights["anchor_summary"] = {
            "num_row_anchors": len(sheet_json_data["structural_anchors"].get("rows", [])),
            "num_col_anchors": len(sheet_json_data["structural_anchors"].get("columns", []))
        }
    return insights

def generate_common_value_map(sheet_json_data, top_n=5, min_len=4):
    """
    Identifies common, non-numeric string values in a sheet's cell data to create a map.

    Args:
        sheet_json_data (dict): The JSON data for a single sheet.
        top_n (int): The number of most common string values to include in the map.
        min_len (int): The minimum length for a string value to be considered.

    Returns:
        dict: A map where keys are placeholders (e.g., "@v1") and values are the
              identified common string values. Returns an empty dict if no suitable
              common strings are found.
    """
    if "cells" not in sheet_json_data or not isinstance(sheet_json_data["cells"], dict):
        return {}

    string_frequencies = {}
    for value_str, refs in sheet_json_data["cells"].items():
        if isinstance(value_str, str) and len(value_str) >= min_len:
            # Attempt to filter out strings that are purely numeric
            is_numeric_string = False
            try:
                # Check if it can be converted to float without error
                float(value_str)
                # Further check: does it look like a number (e.g. "123", "12.34", "-5")
                # This regex checks for optional sign, digits, optional decimal part with digits
                if re.fullmatch(r"[-+]?\d+(\.\d+)?", value_str):
                    is_numeric_string = True
            except ValueError:
                # Not a float, so definitely not a simple numeric string in that sense
                pass

            if not is_numeric_string:
                string_frequencies[value_str] = len(refs)

    if not string_frequencies:
        return {}

    # Sort by frequency (desc) then by string length (desc, longer preferred for same freq), then alphabetically
    sorted_common_strings = sorted(
        string_frequencies.items(),
        key=lambda item: (item[1], len(item[0]), item[0]),
        reverse=True
    )

    value_map = {}
    for i, (value, _freq) in enumerate(sorted_common_strings[:top_n]):
        value_map[f"@v{i+1}"] = value

    return value_map

# Helper functions for chart-to-table linking
def parse_cell_ref(cell_ref):
    """Parses a cell reference string (e.g., 'A1', '$B$2') into 1-based (col, row) indices."""
    if not cell_ref or not isinstance(cell_ref, str):
        return None, None
    try:
        col_letter, row_idx = coordinate_from_string(cell_ref) # Handles '$'
        col_idx = column_index_from_string(col_letter)
        return col_idx, row_idx # 1-based
    except Exception:
        return None, None

def parse_range_ref(range_str):
    """Parses a range string (e.g., 'A1:B2' or 'A1') into 1-based (start_col, start_row, end_col, end_row) indices."""
    if not range_str or not isinstance(range_str, str):
        return None, None, None, None
    try:
        if ':' in range_str:
            start_ref, end_ref = range_str.split(':', 1)
            start_col, start_row = parse_cell_ref(start_ref)
            end_col, end_row = parse_cell_ref(end_ref)
        else: # Single cell range
            start_col, start_row = parse_cell_ref(range_str)
            end_col, end_row = start_col, start_row

        if start_col is None or start_row is None or end_col is None or end_row is None:
             return None, None, None, None
        return start_col, start_row, end_col, end_row
    except Exception:
        return None, None, None, None

def is_cell_within_parsed_range(cell_col_idx, cell_row_idx, r_start_col, r_start_row, r_end_col, r_end_row):
    """Checks if a cell (1-based indices) is within a given parsed range (1-based indices)."""
    if cell_col_idx is None or cell_row_idx is None or \
       r_start_col is None or r_start_row is None or \
       r_end_col is None or r_end_row is None:
        return False
    return (r_start_col <= cell_col_idx <= r_end_col and \
            r_start_row <= cell_row_idx <= r_end_row)

def main():
    """
    Main function to run the Streamlit application.
    Handles file uploads, processing (Excel/CSV), encoding via Spreadsheet_LLM_Encoder,
    and various post-processing steps (table detection, chart extraction, number format parsing,
    sheet metadata extraction, compression insights, common value mapping, chart-table linking),
    and displays the resulting JSON.
    """
    st.title("Spreadsheet to Encoded JSON")
    st.write("Upload your spreadsheet file (Excel or CSV) and get JSON in return.")

    k_value = st.slider("Neighborhood distance (k)", min_value=0, max_value=10, value=2, step=1)
    uploaded_file = st.file_uploader("Choose a spreadsheet file", type=["xlsx", "xls", "csv"])

    if uploaded_file is not None:
        try:
            file_extension = uploaded_file.name.split(".")[-1].lower()
            json_data = None
            processed_file_path_for_postprocessing = None

            if file_extension in ['xlsx', 'xls']:
                with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_extension}") as tmp:
                    tmp.write(uploaded_file.getvalue())
                    processed_file_path_for_postprocessing = tmp.name

                try:
                    json_data = spreadsheet_llm_encode(processed_file_path_for_postprocessing, k=k_value)

                    if json_data and "sheets" in json_data and processed_file_path_for_postprocessing:
                        source_workbook = openpyxl.load_workbook(processed_file_path_for_postprocessing, data_only=True)
                        for sheet_name_wb in source_workbook.sheetnames:
                            if sheet_name_wb in json_data["sheets"]:
                                current_sheet_obj = source_workbook[sheet_name_wb]
                                sheet_data_node = json_data["sheets"][sheet_name_wb]

                                detected_tables_list = detect_tables_in_sheet(current_sheet_obj)
                                if detected_tables_list:
                                    sheet_data_node["detected_tables"] = detected_tables_list

                                chart_list = extract_chart_info_from_sheet(current_sheet_obj)
                                if chart_list:
                                    sheet_data_node["charts"] = chart_list

                                if "format_regions" in sheet_data_node:
                                    parsed_number_formats_map = {}
                                    for fmt_key_json, _ in sheet_data_node["format_regions"].items():
                                        try:
                                            fmt_details = json.loads(fmt_key_json)
                                            number_format_str = fmt_details.get("number_format")
                                            if number_format_str and number_format_str not in parsed_number_formats_map:
                                                parsed_number_formats_map[number_format_str] = parse_number_format_string(number_format_str)
                                        except json.JSONDecodeError:
                                            st.warning(f"Could not parse format key string: {fmt_key_json}")
                                    if parsed_number_formats_map:
                                        sheet_data_node["parsed_number_formats"] = parsed_number_formats_map

                                sheet_metadata = extract_sheet_metadata(current_sheet_obj)
                                sheet_data_node["sheet_level_metadata"] = sheet_metadata

                                compression_analysis_results = analyze_sheet_for_compression_insights(sheet_data_node)
                                sheet_data_node["compression_insights"] = compression_analysis_results

                                common_value_map_results = generate_common_value_map(sheet_data_node)
                                if common_value_map_results:
                                     sheet_data_node["common_value_map"] = common_value_map_results

                                # Chart to Table Linking
                                if "charts" in sheet_data_node and "detected_tables" in sheet_data_node:
                                    for chart_dict in sheet_data_node["charts"]:
                                        if chart_dict.get("anchor"):
                                            chart_anchor_col, chart_anchor_row = parse_cell_ref(chart_dict["anchor"])
                                            if chart_anchor_col and chart_anchor_row:
                                                linked_tables = []
                                                for table_dict in sheet_data_node["detected_tables"]:
                                                    if table_dict.get("full_range"):
                                                        r_start_col, r_start_row, r_end_col, r_end_row = parse_range_ref(table_dict["full_range"])
                                                        if r_start_col and is_cell_within_parsed_range(chart_anchor_col, chart_anchor_row, r_start_col, r_start_row, r_end_col, r_end_row):
                                                            linked_tables.append(table_dict["full_range"])
                                                if linked_tables:
                                                    chart_dict["linked_table_ranges"] = linked_tables
                finally:
                    if processed_file_path_for_postprocessing and os.path.exists(processed_file_path_for_postprocessing):
                        os.remove(processed_file_path_for_postprocessing)

            elif file_extension == 'csv':
                st.write("Processing CSV: Converting to temporary Excel for encoding...")
                df = None
                json_data = None
                try:
                    df = pd.read_csv(uploaded_file)
                except UnicodeDecodeError:
                    st.warning("UTF-8 decoding failed. Trying latin-1...")
                    try:
                        uploaded_file.seek(0) # Reset file pointer
                        df = pd.read_csv(uploaded_file, encoding='latin-1')
                    except UnicodeDecodeError:
                        st.warning("latin-1 decoding failed. Trying iso-8859-1...")
                        try:
                            uploaded_file.seek(0) # Reset file pointer
                            df = pd.read_csv(uploaded_file, encoding='iso-8859-1')
                        except Exception as e:
                            st.error(f"Error reading CSV file: Could not decode file with UTF-8, latin-1, or ISO-8859-1 encoding. Please ensure the file is saved with one of these encodings. Error details: {e}")
                            json_data = None # Explicitly set to None
                    except Exception as e: # Catch other errors during latin-1 read
                        st.error(f"Error reading CSV file with latin-1 encoding: {e}")
                        json_data = None
                except Exception as e: # Catch other errors during initial UTF-8 read
                    st.error(f"Error reading CSV file: {e}")
                    json_data = None

                if df is not None:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = "Sheet1"
                    for col_idx, column_name in enumerate(df.columns, 1):
                        sheet.cell(row=1, column=col_idx, value=str(column_name))
                    for row_idx, row_data_tuple in enumerate(df.itertuples(index=False), 2):
                        for col_idx, cell_value in enumerate(row_data_tuple, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

                    temp_csv_conversion_dir = tempfile.mkdtemp()
                    processed_file_path_for_postprocessing = os.path.join(temp_csv_conversion_dir, "temp_for_csv.xlsx")

                    try:
                        wb.save(processed_file_path_for_postprocessing)
                        json_data = spreadsheet_llm_encode(processed_file_path_for_postprocessing, k=k_value)

                        if json_data and "sheets" in json_data and processed_file_path_for_postprocessing:
                            source_workbook = openpyxl.load_workbook(processed_file_path_for_postprocessing, data_only=True)
                            for sheet_name_wb in source_workbook.sheetnames:
                                 if sheet_name_wb in json_data["sheets"]:
                                    current_sheet_obj = source_workbook[sheet_name_wb]
                                    sheet_data_node = json_data["sheets"][sheet_name_wb]

                                    detected_tables_list = detect_tables_in_sheet(current_sheet_obj)
                                    if detected_tables_list:
                                        sheet_data_node["detected_tables"] = detected_tables_list

                                    chart_list = extract_chart_info_from_sheet(current_sheet_obj)
                                    if chart_list:
                                        sheet_data_node["charts"] = chart_list

                                    if "format_regions" in sheet_data_node:
                                        parsed_number_formats_map = {}
                                        for fmt_key_json, _ in sheet_data_node["format_regions"].items():
                                            try:
                                                fmt_details = json.loads(fmt_key_json)
                                                number_format_str = fmt_details.get("number_format")
                                                if number_format_str and number_format_str not in parsed_number_formats_map:
                                                    parsed_number_formats_map[number_format_str] = parse_number_format_string(number_format_str)
                                            except json.JSONDecodeError:
                                                st.warning(f"Could not parse format key string: {fmt_key_json}")
                                        if parsed_number_formats_map:
                                            sheet_data_node["parsed_number_formats"] = parsed_number_formats_map

                                    sheet_metadata = extract_sheet_metadata(current_sheet_obj)
                                    sheet_data_node["sheet_level_metadata"] = sheet_metadata

                                    compression_analysis_results = analyze_sheet_for_compression_insights(sheet_data_node)
                                    sheet_data_node["compression_insights"] = compression_analysis_results

                                    common_value_map_results = generate_common_value_map(sheet_data_node)
                                    if common_value_map_results:
                                         sheet_data_node["common_value_map"] = common_value_map_results

                                    # Chart to Table Linking (for CSV, charts won't exist from original, but tables might be detected)
                                    if "charts" in sheet_data_node and "detected_tables" in sheet_data_node:
                                        for chart_dict in sheet_data_node["charts"]: # This list will be empty for CSVs
                                            if chart_dict.get("anchor"):
                                                chart_anchor_col, chart_anchor_row = parse_cell_ref(chart_dict["anchor"])
                                                if chart_anchor_col and chart_anchor_row:
                                                    linked_tables = []
                                                    for table_dict in sheet_data_node["detected_tables"]:
                                                        if table_dict.get("full_range"):
                                                            r_start_col, r_start_row, r_end_col, r_end_row = parse_range_ref(table_dict["full_range"])
                                                            if r_start_col and is_cell_within_parsed_range(chart_anchor_col, chart_anchor_row, r_start_col, r_start_row, r_end_col, r_end_row):
                                                                linked_tables.append(table_dict["full_range"])
                                                    if linked_tables:
                                                        chart_dict["linked_table_ranges"] = linked_tables
                    finally:
                        if processed_file_path_for_postprocessing and os.path.exists(processed_file_path_for_postprocessing):
                            os.remove(processed_file_path_for_postprocessing)
                        if 'temp_csv_conversion_dir' in locals() and temp_csv_conversion_dir and os.path.exists(temp_csv_conversion_dir):
                            os.rmdir(temp_csv_conversion_dir)
                # If df is None (all decoding failed), json_data remains None, and this block is skipped.
                # The existing error message for failed processing will be shown later if json_data is still None.

            if json_data:
                json_str = json.dumps(json_data, indent=2)
                st.subheader("JSON Output")
                st.json(json_data)
                st.subheader("Copy JSON")
                st.text_area("Copy this JSON:", value=json_str, height=250)
                st.markdown(get_download_link(json_data), unsafe_allow_html=True)
            else:
                st.error("Failed to process the spreadsheet and generate JSON data.")
            
        except Exception as e:
            st.error(f"Error processing file: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()

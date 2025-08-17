import argparse
import json
import logging
from typing import List, Dict

from evaluation import load_spreadsheet_dataset, BBox
from Spreadsheet_LLM_Encoder import spreadsheet_llm_encode
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TABLE_DETECTION_PROMPT_TEMPLATE = """
INSTRUCTION:
Given an input that is a string denoting data of cells in an Excel spreadsheet. The input spreadsheet contains many tuples, describing the cells with content in the spreadsheet. Each tuple consists of two elements separated by a '|': the cell content and the cell address/region, like (Year|A1), ( |A1) or (IntNum|A1:B3). The content in some cells such as '#,##0'/'d-mmm-yy'/'H:mm:ss',etc., represents the CELL DATA FORMATS of Excel. The content in some cells such as 'IntNum'/'DateData'/'EmailData',etc., represents a category of data with the same format and similar semantics. For example, 'IntNum' represents integer type data, and 'ScientificNum' represents scientific notation type data. 'A1:B3' represents a region in a spreadsheet, from the first row to the third row and from column A to column B. Some cells with empty content in the spreadsheet are not entered. Now you should tell me the range of the table in a format like A2:D5, and the range of the table should only CONTAIN HEADER REGION and the data region. DON’T include the title or comments. Note that there can be more than one table in a string, so you should return all the RANGE. DON’T ADD OTHER WORDS OR EXPLANATION.

INPUT:
[Encoded Spreadsheet]
"""

def bbox_to_range(bbox: BBox) -> str:
    """Converts a BBox tuple to an Excel-style range string."""
    r1, c1, r2, c2 = bbox
    start_cell = f"{get_column_letter(c1)}{r1}"
    end_cell = f"{get_column_letter(c2)}{r2}"
    if start_cell == end_cell:
        return start_cell
    return f"{start_cell}:{end_cell}"

def format_for_finetuning(encoding: Dict, gt_boxes: List[BBox]) -> Dict:
    """
    Formats the encoded spreadsheet and ground truth into a dictionary
    suitable for fine-tuning (e.g., as a single JSONL line).
    """
    # We'll use the first sheet for simplicity in this example
    sheet_name = next(iter(encoding["sheets"]))
    sheet_data = encoding["sheets"][sheet_name]

    prompt_input = json.dumps(sheet_data, ensure_ascii=False)
    prompt = TABLE_DETECTION_PROMPT_TEMPLATE.replace("[Encoded Spreadsheet]", prompt_input)

    # Format the ground truth as the expected LLM response
    gt_ranges = [bbox_to_range(bbox) for bbox in gt_boxes]
    completion = f"[{', '.join([f'\'range\': \'{r}\'' for r in gt_ranges])}]"

    # This format is compatible with many fine-tuning libraries
    return {"prompt": prompt, "completion": completion}


def main(dataset_dir: str, output_path: str, k: int):
    """
    Main function to prepare data for fine-tuning.
    """
    data = load_spreadsheet_dataset(dataset_dir)

    if not data:
        logger.error("No data found in the specified dataset directory.")
        return

    with open(output_path, 'w', encoding='utf-8') as f:
        for item in data:
            logger.info(f"Processing {item['spreadsheet_path']} for fine-tuning...")

            # 1. Encode the spreadsheet
            encoding = spreadsheet_llm_encode(item["spreadsheet_path"], k=k)
            if not encoding or not encoding.get("sheets"):
                logger.warning(f"Skipping {item['spreadsheet_path']} due to encoding error or empty sheets.")
                continue

            # 2. Format for fine-tuning
            ft_record = format_for_finetuning(encoding, item["bboxes"])

            # 3. Write to JSONL file
            f.write(json.dumps(ft_record) + "\n")

    logger.info(f"Fine-tuning data successfully prepared and saved to {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Prepare spreadsheet table detection data for fine-tuning.")
    parser.add_argument("dataset_dir", help="Path to the spreadsheet dataset directory")
    parser.add_argument("output_path", help="Path to save the output JSONL file for fine-tuning")
    parser.add_argument("--k", type=int, default=2, help="Neighborhood distance for structural anchors (default: 2)")
    args = parser.parse_args()
    main(args.dataset_dir, args.output_path, args.k)

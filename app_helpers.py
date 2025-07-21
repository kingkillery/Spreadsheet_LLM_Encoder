"""Helper utilities for Streamlit app."""

import json
import re
from typing import Dict, List, Tuple, Optional
import pandas as pd

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def detect_tables_in_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[dict]:
    """Detect tables in a worksheet using simple heuristics."""
    if not sheet:
        return []

    try:
        min_row, min_col, max_row, max_col = (
            sheet.min_row,
            sheet.min_column,
            sheet.max_row,
            sheet.max_column,
        )
    except Exception:
        return []

    if max_row == 0 or max_col == 0:
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
            return []
        try:
            sheet.calculate_dimension(force=True)
            min_row, min_col, max_row, max_col = (
                sheet.min_row,
                sheet.min_column,
                sheet.max_row,
                sheet.max_column,
            )
            if max_row == 0 or max_col == 0:
                return []
        except Exception:
            return []

    tables = []
    r_idx = min_row
    while r_idx <= max_row:
        num_populated_cells = 0
        num_bold = 0
        num_bottom_border = 0
        num_all_caps = 0
        num_string_cells = 0
        current_min_col = max_col + 1
        current_max_col = min_col - 1

        for c_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                num_populated_cells += 1
                current_min_col = min(current_min_col, c_idx)
                current_max_col = max(current_max_col, c_idx)

                if cell.font and cell.font.bold:
                    num_bold += 1
                if (
                    cell.border
                    and cell.border.bottom
                    and cell.border.bottom.style is not None
                    and cell.border.bottom.style != "none"
                ):
                    num_bottom_border += 1
                if isinstance(cell.value, str):
                    num_string_cells += 1
                    if cell.value.isupper():
                        num_all_caps += 1

        is_header = False
        if num_populated_cells > 1:
            if (
                num_bold / num_populated_cells > 0.5
                or num_bottom_border / num_populated_cells > 0.5
                or (
                    num_string_cells == num_populated_cells
                    and num_string_cells > 0
                    and num_all_caps / num_string_cells > 0.5
                )
            ):
                is_header = True
        elif num_populated_cells == 1:
            if num_bold == 1 or num_bottom_border == 1 or (
                num_string_cells == 1 and num_all_caps == 1
            ):
                is_header = True

        if is_header:
            header_row_idx = r_idx
            header_min_col = current_min_col
            header_max_col = current_max_col
            table_min_col = header_min_col
            table_max_col = header_max_col
            data_end_row_idx = header_row_idx

            for data_r_idx in range(header_row_idx + 1, max_row + 1):
                data_populated = 0
                data_min_col = max_col + 1
                data_max_col = min_col - 1

                for c_idx_data in range(min_col, max_col + 1):
                    cell_data = sheet.cell(row=data_r_idx, column=c_idx_data)
                    if cell_data.value is not None and str(cell_data.value).strip() != "":
                        data_populated += 1
                        data_min_col = min(data_min_col, c_idx_data)
                        data_max_col = max(data_max_col, c_idx_data)

                if data_populated == 0:
                    break

                if (
                    data_populated > 0
                    and data_max_col >= table_min_col
                    and data_min_col <= table_max_col
                ):
                    data_end_row_idx = data_r_idx
                    table_min_col = min(table_min_col, data_min_col)
                    table_max_col = max(table_max_col, data_max_col)
                else:
                    break

            if data_end_row_idx >= header_row_idx:
                header_range = f"{get_column_letter(table_min_col)}{header_row_idx}:{get_column_letter(table_max_col)}{header_row_idx}"
                data_range = None
                if data_end_row_idx > header_row_idx:
                    data_range = f"{get_column_letter(table_min_col)}{header_row_idx + 1}:{get_column_letter(table_max_col)}{data_end_row_idx}"
                full_range = f"{get_column_letter(table_min_col)}{header_row_idx}:{get_column_letter(table_max_col)}{data_end_row_idx}"

                tables.append(
                    {
                        "full_range": full_range,
                        "header_range": header_range,
                        "data_range": data_range,
                        "detection_method": "improved_heuristic_v1",
                    }
                )
                r_idx = data_end_row_idx + 1
                continue
        r_idx += 1
    return tables


def extract_chart_info_from_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[dict]:
    """Extract basic information about charts in a worksheet."""
    chart_info_list = []
    if not hasattr(sheet, "_charts") or not sheet._charts:
        return chart_info_list

    for chart in sheet._charts:
        title_str = None
        if chart.title:
            if (
                hasattr(chart.title, "tx")
                and hasattr(chart.title.tx, "rich")
                and hasattr(chart.title.tx.rich, "p")
                and len(chart.title.tx.rich.p) > 0
                and hasattr(chart.title.tx.rich.p[0], "r")
                and len(chart.title.tx.rich.p[0].r) > 0
                and hasattr(chart.title.tx.rich.p[0].r[0], "t")
            ):
                title_str = chart.title.tx.rich.p[0].r[0].t
            elif isinstance(chart.title, str):
                title_str = chart.title

        chart_type_str = chart.type if chart.type else None
        anchor_cell_str = None

        try:
            if hasattr(chart, "anchor"):
                anchor = chart.anchor
                if hasattr(anchor, "twoCellAnchor"):
                    _from = anchor.twoCellAnchor._from
                    anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
                elif hasattr(anchor, "oneCellAnchor"):
                    _from = getattr(anchor.oneCellAnchor, "_from", None)
                    if _from and hasattr(_from, "col") and hasattr(_from, "row"):
                        anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
                    elif hasattr(anchor.oneCellAnchor, "cell") and hasattr(anchor.oneCellAnchor.cell, "coordinate"):
                        anchor_cell_str = anchor.oneCellAnchor.cell.coordinate
                elif hasattr(anchor, "_from") and hasattr(anchor._from, "col") and hasattr(anchor._from, "row"):
                    _from = anchor._from
                    anchor_cell_str = f"{get_column_letter(_from.col + 1)}{_from.row + 1}"
        except Exception:
            pass

        chart_info_list.append({"title": title_str, "type": chart_type_str, "anchor": anchor_cell_str})
    return chart_info_list


def parse_number_format_string(format_str: str) -> Dict[str, Optional[str]]:
    """Parse an Excel number format string."""
    if format_str is None:
        return {"type": "none", "original": None}
    if not isinstance(format_str, str):
        return {"type": "invalid", "original": str(format_str)}

    original_format_str = format_str
    format_str_lower = format_str.lower()

    if format_str_lower == "general":
        return {"type": "general", "original": original_format_str}
    if format_str_lower in {"@", "text"}:
        return {"type": "text", "original": original_format_str}

    if "%" in format_str:
        decimals = 0
        match = re.search(r"\.([0#]+)(?=[^%]*%)", format_str)
        if match:
            decimals = len(match.group(1))
        return {"type": "percentage", "decimals": decimals, "original": original_format_str}

    currency_match = re.search(r"(?<!\[[^\]]*)([$€£¥])", format_str)
    if currency_match:
        symbol_map = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "JPY"}
        found_symbol_char = currency_match.group(1)
        currency_code = symbol_map.get(found_symbol_char, found_symbol_char)

        decimals = 0
        match = re.search(r"\.([0#]+)", format_str)
        if match:
            decimals = len(match.group(1))
        return {
            "type": "currency",
            "symbol": currency_code,
            "decimals": decimals,
            "original": original_format_str,
        }

    has_date_chars = any(c in format_str_lower for c in ["y", "d"]) or "mmm" in format_str_lower
    has_time_chars = any(c in format_str_lower for c in ["h", "s", "am/pm", "ss"])

    is_date = False
    is_time = False

    if has_date_chars:
        is_date = True
    if has_time_chars:
        is_time = True

    if "m" in format_str_lower and not is_date and not is_time:
        if re.fullmatch(r"m{1,5}", format_str_lower):
            is_date = True
        elif ("h" in format_str_lower or "s" in format_str_lower) and "m" in format_str_lower:
            is_time = True

    if is_date and is_time:
        return {"type": "datetime", "format_tokens": original_format_str, "original": original_format_str}
    if is_date:
        return {"type": "date", "format_tokens": original_format_str, "original": original_format_str}
    if is_time:
        return {"type": "time", "format_tokens": original_format_str, "original": original_format_str}

    if "e+" in format_str_lower or "e-" in format_str_lower:
        decimals = 0
        match = re.search(r"\.([0#]+)(?=E)", format_str, re.IGNORECASE)
        if match:
            decimals = len(match.group(1))
        return {"type": "scientific", "decimals": decimals, "original": original_format_str}

    if "/" in format_str and "?" in format_str:
        return {"type": "fraction", "original": original_format_str}

    if re.fullmatch(r"[#0,]+(\.[#0]+)?", format_str) or re.fullmatch(r"[#0,]+", format_str):
        decimals = 0
        if "." in format_str:
            match = re.search(r"\.([0#]+)", format_str)
            if match:
                decimals = len(match.group(1))
        has_comma = "," in format_str.split(".")[0]
        return {
            "type": "number",
            "decimals": decimals,
            "thousands_separator": has_comma,
            "original": original_format_str,
        }

    return {"type": "unknown", "original": original_format_str}


def extract_sheet_metadata(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Optional[str]]:
    """Extract sheet level metadata."""
    visibility = sheet.sheet_state if hasattr(sheet, "sheet_state") else "visible"
    is_protected = False
    if hasattr(sheet, "protection") and hasattr(sheet.protection, "sheet"):
        is_protected = sheet.protection.sheet

    tab_color_str = None
    if (
        hasattr(sheet, "sheet_properties")
        and hasattr(sheet.sheet_properties, "tabColor")
        and sheet.sheet_properties.tabColor
    ):
        tab_color_obj = sheet.sheet_properties.tabColor
        if hasattr(tab_color_obj, "rgb") and tab_color_obj.rgb:
            tab_color_str = str(tab_color_obj.rgb)
        elif hasattr(tab_color_obj, "indexed") and tab_color_obj.indexed > 0:
            tab_color_str = f"indexed_{tab_color_obj.indexed}"
        elif hasattr(tab_color_obj, "theme") and hasattr(tab_color_obj, "tint"):
            tab_color_str = f"theme_{tab_color_obj.theme}_tint_{tab_color_obj.tint}"

    return {"visibility": visibility, "is_protected": is_protected, "tab_color": tab_color_str}


def analyze_sheet_for_compression_insights(sheet_json_data: dict) -> Dict[str, dict]:
    """Analyze sheet JSON data for potential compression insights."""
    insights: Dict[str, dict] = {}

    if "format_regions" in sheet_json_data and isinstance(sheet_json_data["format_regions"], dict):
        format_keys_json_strings = list(sheet_json_data["format_regions"].keys())
        num_unique_formats_overall = len(format_keys_json_strings)
        potential_redundancy_groups = []
        base_format_groups = {}

        for fmt_key_str in format_keys_json_strings:
            try:
                fmt_details = json.loads(fmt_key_str)
                core_props_dict = {
                    "font": fmt_details.get("font"),
                    "alignment": fmt_details.get("alignment"),
                    "border": fmt_details.get("border"),
                    "fill": fmt_details.get("fill"),
                }

                def make_hashable(obj):
                    if isinstance(obj, dict):
                        return tuple(sorted((k, make_hashable(v)) for k, v in obj.items()))
                    if isinstance(obj, list):
                        return tuple(make_hashable(elem) for elem in obj)
                    return obj

                base_key = make_hashable(core_props_dict)
                if base_key not in base_format_groups:
                    base_format_groups[base_key] = []
                base_format_groups[base_key].append(fmt_details.get("number_format", "General"))
            except json.JSONDecodeError:
                continue

        for base_fmt_tuple, number_format_list in base_format_groups.items():
            if len(number_format_list) > 1:
                unique_number_formats_in_group = sorted(set(number_format_list))
                if len(unique_number_formats_in_group) > 1:
                    potential_redundancy_groups.append(
                        {
                            "base_format_properties_hash": str(base_fmt_tuple),
                            "differing_number_formats": unique_number_formats_in_group,
                            "count_of_full_format_keys_in_group": len(number_format_list),
                        }
                    )
        insights["format_analysis"] = {
            "num_unique_formats_overall": num_unique_formats_overall,
            "num_base_format_groups": len(base_format_groups),
            "potential_redundancy_groups": sorted(
                potential_redundancy_groups,
                key=lambda x: x["count_of_full_format_keys_in_group"],
                reverse=True,
            ),
        }

    if "compressed_cells" in sheet_json_data and isinstance(sheet_json_data["compressed_cells"], dict):
        value_counts = {value: len(refs) for value, refs in sheet_json_data["compressed_cells"].items()}
        sorted_values_by_freq = sorted(value_counts.items(), key=lambda item: (item[1], str(item[0])), reverse=True)
        insights["value_frequency"] = {
            "num_unique_values": len(sorted_values_by_freq),
            "top_10_frequent_values": sorted_values_by_freq[:10],
        }
        total_cell_references = sum(len(refs) for refs in sheet_json_data["compressed_cells"].values())
        insights["total_cell_references_in_index"] = total_cell_references

    if "structural_anchors" in sheet_json_data and isinstance(sheet_json_data["structural_anchors"], dict):
        insights["anchor_summary"] = {
            "num_row_anchors": len(sheet_json_data["structural_anchors"].get("rows", [])),
            "num_col_anchors": len(sheet_json_data["structural_anchors"].get("columns", [])),
        }
    return insights


def generate_common_value_map(sheet_json_data: dict, top_n: int = 5, min_len: int = 4) -> Dict[str, str]:
    """Identify common non-numeric string values in a sheet."""
    if "compressed_cells" not in sheet_json_data or not isinstance(sheet_json_data["compressed_cells"], dict):
        return {}

    string_frequencies = {}
    for value_str, refs in sheet_json_data["compressed_cells"].items():
        if isinstance(value_str, str) and len(value_str) >= min_len:
            is_numeric_string = False
            try:
                float(value_str)
                if re.fullmatch(r"[-+]?\d+(\.\d+)?", value_str):
                    is_numeric_string = True
            except ValueError:
                pass

            if not is_numeric_string:
                string_frequencies[value_str] = len(refs)

    if not string_frequencies:
        return {}

    sorted_common_strings = sorted(string_frequencies.items(), key=lambda item: (item[1], len(item[0]), item[0]), reverse=True)

    value_map = {}
    for i, (value, _freq) in enumerate(sorted_common_strings[:top_n]):
        value_map[f"@v{i + 1}"] = value

    return value_map


# Helper functions for chart-to-table linking

def parse_cell_ref(cell_ref: str) -> Tuple[Optional[int], Optional[int]]:
    """Parse a cell reference like 'A1' into (col, row) indices."""
    if not cell_ref or not isinstance(cell_ref, str):
        return None, None
    try:
        col_letter, row_idx = coordinate_from_string(cell_ref)
        col_idx = column_index_from_string(col_letter)
        return col_idx, row_idx
    except Exception:
        return None, None


def parse_range_ref(range_str: str) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Parse a range string like 'A1:B2'."""
    if not range_str or not isinstance(range_str, str):
        return None, None, None, None
    try:
        if ":" in range_str:
            start_ref, end_ref = range_str.split(":", 1)
            start_col, start_row = parse_cell_ref(start_ref)
            end_col, end_row = parse_cell_ref(end_ref)
        else:
            start_col, start_row = parse_cell_ref(range_str)
            end_col, end_row = start_col, start_row

        if None in (start_col, start_row, end_col, end_row):
            return None, None, None, None
        return start_col, start_row, end_col, end_row
    except Exception:
        return None, None, None, None


def is_cell_within_parsed_range(
    cell_col_idx: int,
    cell_row_idx: int,
    r_start_col: int,
    r_start_row: int,
    r_end_col: int,
    r_end_row: int,
) -> bool:
    """Check if a cell is within a given parsed range."""
    if None in (cell_col_idx, cell_row_idx, r_start_col, r_start_row, r_end_col, r_end_row):
        return False
    return r_start_col <= cell_col_idx <= r_end_col and r_start_row <= cell_row_idx <= r_end_row


def read_csv_with_multiple_encodings(file) -> Optional[pd.DataFrame]:
    """Try reading a CSV file with multiple encodings."""
    encodings = ["utf-8", "latin-1", "iso-8859-1"]
    for enc in encodings:
        try:
            file.seek(0)
            return pd.read_csv(file, encoding=enc)
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    return None
